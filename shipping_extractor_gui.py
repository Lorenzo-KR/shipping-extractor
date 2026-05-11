"""
Shipping Document Extractor - GUI
===================================
사용법: python shipping_extractor_gui.py
필요 라이브러리: pip install pdfplumber openpyxl
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import os
import sys
import glob
from datetime import datetime

# ── 파싱/Excel 로직 (extract_shipping_docs.py 와 동일) ──────────────────

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

def parse_invoice_page(text):
    invoice_no, invoice_date = '', ''
    m = re.search(r'Invoice No\.\s*:\s*(\S+)', text)
    if m: invoice_no = m.group(1)
    m = re.search(r'Date\s*:\s*([\w.]+\s+\d+,\s+\d+)', text)
    if m: invoice_date = m.group(1)
    items = {}
    item_pattern = re.compile(
        r'(\d{10})\s+(\d+)\s+(\S+)\s+([\d,]+)\s+([\d,.]+)\s+([\d,.]+)\s+(\w+)\s+(\w+)\s+(\w+)\s+(\S+)(?:\s+(.+))?'
    )
    for m in item_pattern.finditer(text):
        delivery_no = m.group(1)
        items[delivery_no] = {
            'Invoice_No':     invoice_no,
            'Invoice_Date':   invoice_date,
            'Mat_Code':       m.group(2),
            'Unit_Price_USD': float(m.group(5).replace(',', '')),
            'Amt_USD':        float(m.group(6).replace(',', '')),
        }
    return items

def parse_packing_detail_page(table):
    rows = []
    if not table: return rows
    last_pallet_no = ''
    last_pallet_wt = ''
    for row in table:
        first = str(row[0] or '')
        if 'Pallet' in first or first == 'L' or 'Grand Total' in first: continue
        if not row[1]: continue
        pallet_no = row[0] if row[0] else last_pallet_no
        pallet_wt = row[7] if row[7] else last_pallet_wt
        last_pallet_no = pallet_no
        last_pallet_wt = pallet_wt
        rows.append({
            'Pallet_No':    pallet_no,
            'Delivery_No':  row[1],
            'CT_No':        row[2],
            'Description':  row[3],
            'Qty_EA':       _to_int(row[4]),
            'Net_Wt_Kg':    _to_float(row[5]),
            'Gross_Wt_Kg':  _to_float(row[6]),
            'Pallet_Wt_Kg': pallet_wt,
            'LWH_cm':       f"{_to_int(row[8])} x {_to_int(row[9])} x {_to_int(row[10])}",
            'COO':          row[11],
            'Remarks':      row[12],
        })
    return rows

def _to_float(val):
    try: return float(str(val).replace(',', ''))
    except: return val

def _to_int(val):
    try: return int(str(val).replace(',', ''))
    except: return val

def process_pdf(pdf_path):
    invoice_data, detail_rows = {}, []
    source_file = os.path.basename(pdf_path)
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                tables = page.extract_tables()
                if 'Commercial Invoice' in text:
                    invoice_data.update(parse_invoice_page(text))
                elif text.strip().startswith('Packing Details') and tables:
                    rows = parse_packing_detail_page(tables[0])
                    for r in rows: r['Source_File'] = source_file
                    detail_rows.extend(rows)
    except Exception as e:
        return invoice_data, detail_rows, str(e)
    return invoice_data, detail_rows, None

def process_all(pdf_paths, log_fn=None):
    all_rows = []
    total = len(pdf_paths)
    for i, path in enumerate(pdf_paths, 1):
        fname = os.path.basename(path)
        if log_fn: log_fn(f"[{i}/{total}] {fname}")
        invoice_data, detail_rows, err = process_pdf(path)
        if err and log_fn: log_fn(f"  ⚠️  오류: {err}")
        seen_delivery = set()
        for row in detail_rows:
            delivery_no = row.get('Delivery_No', '')
            inv = invoice_data.get(delivery_no, {})
            # 같은 Delivery No.의 첫 번째 행에만 U/P, Amt 표시
            is_first = delivery_no not in seen_delivery
            seen_delivery.add(delivery_no)
            row.update({
                'Invoice_No':     inv.get('Invoice_No', ''),
                'Invoice_Date':   inv.get('Invoice_Date', ''),
                'Mat_Code':       inv.get('Mat_Code', ''),
                'Unit_Price_USD': inv.get('Unit_Price_USD', '') if is_first else '',
                'Amt_USD':        inv.get('Amt_USD', '') if is_first else '',
            })
            all_rows.append(row)
    return all_rows

COLUMNS = [
    ('Source_File',    'Source File',      25),
    ('Invoice_No',     'Invoice No.',       15),
    ('Invoice_Date',   'Invoice Date',      15),
    ('Delivery_No',    'Delivery No.',      13),
    ('CT_No',          'C/T No.',           10),
    ('Description',    'Description',       20),
    ('Qty_EA',         'Qty (EA)',           10),
    ('Unit_Price_USD', 'U/P (USD)',          12),
    ('Amt_USD',        'Amt (USD)',          13),
    ('Net_Wt_Kg',      'Net Wt (Kg)',       12),
    ('Gross_Wt_Kg',    'Gross Wt (Kg)',     13),
    ('Pallet_Wt_Kg',   'Pallet Wt (Kg)',    14),
    ('LWH_cm',         'L x W x H (cm)',    16),
    ('COO',            'COO',                7),
    ('Remarks',        'Remarks',           15),
    ('Pallet_No',      'Pallet No.',        10),
    ('Mat_Code',       'Mat. Code',         12),
]

def write_excel(rows, out_path):
    HDR_FILL = PatternFill('solid', fgColor='D9D9D9')
    HDR_FONT = Font(bold=True, color='000000', size=10, name='Calibri')
    DATA_FONT = Font(size=10, name='Calibri')
    ALT_FILL = PatternFill('solid', fgColor='EEF2F7')
    BORDER = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Shipping Data'
    for col_idx, (_, header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'
    for row_idx, row in enumerate(rows, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, (key, _, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ''))
            cell.font = DATA_FONT; cell.border = BORDER
            cell.alignment = Alignment(vertical='center')
            if fill: cell.fill = fill
    num_cols = {'U/P (USD)': '#,##0.00', 'Amt (USD)': '#,##0.00',
                'Net Wt (Kg)': '#,##0.000', 'Gross Wt (Kg)': '#,##0.000', 'Qty (EA)': '#,##0'}
    for col_idx, (_, header, _) in enumerate(COLUMNS, 1):
        if header in num_cols:
            for r in range(2, len(rows) + 2):
                ws.cell(row=r, column=col_idx).number_format = num_cols[header]
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = '추출 일시';  ws2['B1'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws2['A2'] = '총 행 수';   ws2['B2'] = len(rows)
    ws2['A3'] = '총 수량 (EA)'; ws2['B3'] = sum(r.get('Qty_EA', 0) for r in rows if isinstance(r.get('Qty_EA'), (int, float)))
    ws2['A4'] = '총 금액 (USD)'; ws2['B4'] = round(sum(r.get('Amt_USD', 0) for r in rows if isinstance(r.get('Amt_USD'), (int, float))), 2)
    ws2['B4'].number_format = '#,##0.00'
    wb.save(out_path)


# ── GUI ─────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('Shipping Document Extractor')
        self.resizable(True, True)
        self.configure(bg='#F5F7FA')
        self._build_ui()
        self._center()

    def _center(self):
        self.update_idletasks()
        w, h = 580, 580
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f'{w}x{h}+{(sw-w)//2}+{(sh-h)//2}')
        self.minsize(500, 520)

    def _build_ui(self):
        PAD = 20
        BG = '#F5F7FA'
        ACCENT = '#1F3864'
        BTN_FG = 'white'

        # 타이틀
        tk.Label(self, text='Shipping Document Extractor',
                 font=('Segoe UI', 14, 'bold'), bg=BG, fg=ACCENT).pack(pady=(PAD, 4))
        tk.Label(self, text='Commercial Invoice + Packing Detail PDF → Excel',
                 font=('Segoe UI', 9), bg=BG, fg='#666').pack(pady=(0, PAD))

        # PDF 폴더 선택
        frm1 = tk.LabelFrame(self, text=' 📁  PDF 폴더 ', font=('Segoe UI', 9, 'bold'),
                              bg=BG, fg=ACCENT, bd=1, relief='groove', padx=12, pady=10)
        frm1.pack(fill='x', padx=PAD, pady=(0, 10))

        self.folder_var = tk.StringVar(value='폴더를 선택하세요')
        tk.Entry(frm1, textvariable=self.folder_var, font=('Segoe UI', 9),
                 state='readonly', width=44, bg='white').pack(side='left', padx=(0, 8))
        tk.Button(frm1, text='찾아보기', font=('Segoe UI', 9), bg=ACCENT, fg=BTN_FG,
                  relief='flat', padx=10, command=self._browse_folder).pack(side='left')

        # 저장 위치
        frm2 = tk.LabelFrame(self, text=' 💾  저장 위치 ', font=('Segoe UI', 9, 'bold'),
                              bg=BG, fg=ACCENT, bd=1, relief='groove', padx=12, pady=10)
        frm2.pack(fill='x', padx=PAD, pady=(0, 10))

        self.out_var = tk.StringVar(value='PDF 폴더와 같은 위치에 저장')
        tk.Entry(frm2, textvariable=self.out_var, font=('Segoe UI', 9),
                 state='readonly', width=44, bg='white').pack(side='left', padx=(0, 8))
        tk.Button(frm2, text='변경', font=('Segoe UI', 9), bg='#555', fg=BTN_FG,
                  relief='flat', padx=10, command=self._browse_out).pack(side='left')

        # 로그
        frm3 = tk.LabelFrame(self, text=' 📋  처리 로그 ', font=('Segoe UI', 9, 'bold'),
                              bg=BG, fg=ACCENT, bd=1, relief='groove', padx=8, pady=8)
        frm3.pack(fill='both', expand=True, padx=PAD, pady=(0, 10))

        self.log_text = tk.Text(frm3, height=10, font=('Consolas', 9), bg='#1E1E1E',
                                fg='#D4D4D4', relief='flat', state='disabled', wrap='word')
        sb = tk.Scrollbar(frm3, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        self.log_text.pack(fill='both', expand=True)

        # 실행 버튼
        self.run_btn = tk.Button(self, text='▶  추출 시작', font=('Segoe UI', 11, 'bold'),
                                 bg='#2E7D32', fg=BTN_FG, relief='flat', height=2,
                                 activebackground='#1B5E20', command=self._run)
        self.run_btn.pack(fill='x', padx=PAD, pady=(0, PAD))

        self._out_dir = None  # 별도 지정 저장 경로

    def _browse_folder(self):
        path = filedialog.askdirectory(title='PDF 파일이 있는 폴더 선택')
        if path:
            self.folder_var.set(path)

    def _browse_out(self):
        path = filedialog.askdirectory(title='Excel 저장 위치 선택')
        if path:
            self._out_dir = path
            self.out_var.set(path)

    def _log(self, msg):
        self.log_text.configure(state='normal')
        self.log_text.insert('end', msg + '\n')
        self.log_text.see('end')
        self.log_text.configure(state='disabled')
        self.update_idletasks()

    def _run(self):
        folder = self.folder_var.get()
        if not folder or folder == '폴더를 선택하세요':
            messagebox.showwarning('폴더 미선택', 'PDF 폴더를 먼저 선택하세요.')
            return

        pdf_paths = sorted(glob.glob(os.path.join(folder, '*.pdf')) +
                           glob.glob(os.path.join(folder, '*.PDF')))
        if not pdf_paths:
            messagebox.showwarning('PDF 없음', f'선택한 폴더에 PDF 파일이 없습니다.\n{folder}')
            return

        self.run_btn.configure(state='disabled', text='처리 중...', bg='#888')
        self.log_text.configure(state='normal')
        self.log_text.delete('1.0', 'end')
        self.log_text.configure(state='disabled')

        def worker():
            self._log(f'📄 PDF {len(pdf_paths)}개 처리 시작\n')
            rows = process_all(pdf_paths, log_fn=self._log)

            if not rows:
                self._log('\n⚠️  추출된 데이터가 없습니다.')
                self.run_btn.configure(state='normal', text='▶  추출 시작', bg='#2E7D32')
                return

            out_dir = self._out_dir if self._out_dir else folder
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            out_path = os.path.join(out_dir, f'shipping_extracted_{timestamp}.xlsx')

            try:
                write_excel(rows, out_path)
                self._log(f'\n✅ 완료!  총 {len(rows)}행 추출')
                self._log(f'📂 {out_path}')
                if messagebox.askyesno('완료', f'✅ 추출 완료!\n\n총 {len(rows)}행\n\nExcel 파일을 열까요?'):
                    os.startfile(out_path)
            except Exception as e:
                self._log(f'\n❌ 저장 오류: {e}')

            self.run_btn.configure(state='normal', text='▶  추출 시작', bg='#2E7D32')

        threading.Thread(target=worker, daemon=True).start()


if __name__ == '__main__':
    app = App()
    app.mainloop()
